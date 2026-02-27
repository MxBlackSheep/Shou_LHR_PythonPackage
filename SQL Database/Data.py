import os
import re
import sys
from datetime import datetime
from typing import Dict, List, Optional, Set, Tuple

import pandas as pd
import pyodbc


LOG_DIR = r"C:\Python Log"
os.makedirs(LOG_DIR, exist_ok=True)

SCRIPT_NAME = os.path.splitext(os.path.basename(sys.argv[0]))[0]
TS = datetime.now().strftime("%Y%m%d_%H%M%S")
LOG_FILE = os.path.join(LOG_DIR, f"{SCRIPT_NAME}_{TS}.log")

OUT_DIR = r"C:\EvoExports"
os.makedirs(OUT_DIR, exist_ok=True)


def log(msg: str) -> None:
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(f"[{now}] {msg}\n")


def safe_print(msg: str) -> None:
    try:
        print(msg, flush=True)
    except Exception:
        pass


def connect() -> pyodbc.Connection:
    return pyodbc.connect(
        "DRIVER={ODBC Driver 11 for SQL Server};"
        "SERVER=LOCALHOST\\HAMILTON;"
        "DATABASE=EvoYeast;"
        "Trusted_Connection=yes;"
    )


def fetch_df(conn: pyodbc.Connection, sql: str, params: Tuple = ()) -> pd.DataFrame:
    return pd.read_sql(sql, conn, params=params)


def table_exists(conn: pyodbc.Connection, schema: str, table: str) -> bool:
    sql = """
    SELECT 1
    FROM INFORMATION_SCHEMA.TABLES
    WHERE TABLE_SCHEMA = ? AND TABLE_NAME = ?
    """
    df = fetch_df(conn, sql, (schema, table))
    return not df.empty


def get_table_columns(conn: pyodbc.Connection, schema: str, table: str) -> List[str]:
    sql = """
    SELECT COLUMN_NAME
    FROM INFORMATION_SCHEMA.COLUMNS
    WHERE TABLE_SCHEMA = ? AND TABLE_NAME = ?
    ORDER BY ORDINAL_POSITION
    """
    df = fetch_df(conn, sql, (schema, table))
    if df.empty:
        return []
    return df["COLUMN_NAME"].astype(str).tolist()


def select_existing(conn: pyodbc.Connection, schema: str, table: str, desired: List[str]) -> List[str]:
    cols = get_table_columns(conn, schema, table)
    if not cols:
        return []
    cols_map = {c.lower(): c for c in cols}
    out: List[str] = []
    for d in desired:
        real = cols_map.get(d.lower())
        if real:
            out.append(real)
    return out


def get_ancestor_plate_ids(conn: pyodbc.Connection, experiment_id: int) -> List[int]:
    sql = """
    SELECT DISTINCT PlateID
    FROM dbo.AncestPlatesInExperiments
    WHERE ExperimentID = ?
    """
    df = fetch_df(conn, sql, (experiment_id,))
    if df.empty or "PlateID" not in df.columns:
        return []
    return sorted(df["PlateID"].dropna().astype(int).unique().tolist())


def get_all_plate_ids(conn: pyodbc.Connection, ancestor_plate_ids: List[int]) -> List[int]:
    all_plates = set()
    for anc in ancestor_plate_ids:
        anc = int(anc)
        all_plates.add(anc)

        df = fetch_df(conn, "SELECT DISTINCT DescPlateID FROM dbo.Descendants(?)", (anc,))
        if not df.empty and "DescPlateID" in df.columns:
            for x in df["DescPlateID"].dropna().astype(int).tolist():
                all_plates.add(int(x))

    return sorted(all_plates)


def get_cultures(conn: pyodbc.Connection, plate_ids: List[int]) -> pd.DataFrame:
    if not plate_ids:
        return pd.DataFrame(columns=["CultureID", "PlateID", "WellID"])

    placeholders = ",".join(["?"] * len(plate_ids))
    sql = f"""
    SELECT CultureID, PlateID, WellID
    FROM dbo.Cultures
    WHERE PlateID IN ({placeholders})
    """
    df = fetch_df(conn, sql, tuple(plate_ids))
    if df.empty:
        return pd.DataFrame(columns=["CultureID", "PlateID", "WellID"])

    df["CultureID"] = df["CultureID"].astype(int)
    df["PlateID"] = df["PlateID"].astype(int)
    df["WellID"] = df["WellID"].astype(str)
    return df


def get_cells_count_from_import_plate_pattern(conn: pyodbc.Connection, plate_ids: List[int]) -> int:
    candidates = [("dbo", "ImportPlatePattern"), ("dbo", "ImportPlatePatterns")]
    found = None
    for schema, table in candidates:
        if table_exists(conn, schema, table):
            found = (schema, table)
            break
    if found is None:
        raise RuntimeError("Cannot find dbo.ImportPlatePattern table")

    schema, table = found
    desired = ["PlateID", "WellAssign"]
    existing = select_existing(conn, schema, table, desired)
    missing = [x for x in desired if x.lower() not in {c.lower() for c in existing}]
    if missing:
        raise RuntimeError(f"{schema}.{table} missing required columns: {missing}")

    placeholders = ",".join(["?"] * len(plate_ids))
    sql = f"""
    SELECT COUNT(*) AS n
    FROM {schema}.{table}
    WHERE PlateID IN ({placeholders})
      AND LOWER(WellAssign) = 'cells'
    """
    df = fetch_df(conn, sql, tuple(plate_ids))
    if df.empty:
        return 0
    return int(df.loc[0, "n"])


def get_propagation_edges(conn: pyodbc.Connection) -> pd.DataFrame:
    cols = get_table_columns(conn, "dbo", "Propagation")
    if not cols:
        raise RuntimeError("dbo.Propagation not found or has no columns")

    cols_lower = [c.lower() for c in cols]

    def pick(candidates: List[str]) -> str:
        for cand in candidates:
            if cand in cols_lower:
                return cols[cols_lower.index(cand)]
        raise RuntimeError(f"Cannot find any of {candidates} in dbo.Propagation columns: {cols}")

    parent_col = pick(["parentcultureid", "parent_culture_id", "parentid", "parent"])
    child_col = pick(["chldcultureid", "childcultureid", "child_culture_id", "chldid", "childid", "child"])

    sql = f"""
    SELECT [{parent_col}] AS ParentCultureID,
           [{child_col}] AS ChildCultureID
    FROM dbo.Propagation
    """
    df = fetch_df(conn, sql)
    if df.empty:
        return pd.DataFrame(columns=["ParentCultureID", "ChildCultureID"])

    df = df.dropna(subset=["ParentCultureID", "ChildCultureID"]).copy()
    df["ParentCultureID"] = pd.to_numeric(df["ParentCultureID"], errors="coerce")
    df["ChildCultureID"] = pd.to_numeric(df["ChildCultureID"], errors="coerce")
    df = df.dropna(subset=["ParentCultureID", "ChildCultureID"]).copy()
    df["ParentCultureID"] = df["ParentCultureID"].astype(int)
    df["ChildCultureID"] = df["ChildCultureID"].astype(int)
    return df


def get_history_table(
    conn: pyodbc.Connection,
    schema: str,
    table: str,
    culture_ids: List[int],
) -> pd.DataFrame:
    if not culture_ids:
        return pd.DataFrame(columns=["CultureID", "TimeStamp"])

    desired = [
        "CultureID",
        "Iteration",
        "TimeStamp",
        "Converted_OD",
        "converted_od",
        "FlEx482Em510",
        "FlEx587Em611",
        "OD_FlEx482Em510",
        "OD_FlEx587Em611",
    ]
    existing = select_existing(conn, schema, table, desired)
    if not existing:
        return pd.DataFrame(columns=["CultureID", "TimeStamp"])

    col_sql = ", ".join([f"[{c}]" for c in existing])
    placeholders = ",".join(["?"] * len(culture_ids))
    sql = f"""
    SELECT {col_sql}
    FROM {schema}.{table}
    WHERE CultureID IN ({placeholders})
    """
    df = fetch_df(conn, sql, tuple(culture_ids))
    if df.empty:
        return pd.DataFrame(columns=existing)

    df["CultureID"] = pd.to_numeric(df["CultureID"], errors="coerce").astype("Int64")
    df = df.dropna(subset=["CultureID"]).copy()
    df["CultureID"] = df["CultureID"].astype(int)

    if "Iteration" in df.columns:
        df["Iteration"] = pd.to_numeric(df["Iteration"], errors="coerce")

    df["TimeStamp"] = pd.to_datetime(df["TimeStamp"])

    for c in ["Converted_OD", "converted_od", "FlEx482Em510", "FlEx587Em611", "OD_FlEx482Em510", "OD_FlEx587Em611"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")

    if "Converted_OD" not in df.columns and "converted_od" in df.columns:
        df["Converted_OD"] = df["converted_od"]

    return df


def get_cultures_history(conn: pyodbc.Connection, culture_ids: List[int]) -> pd.DataFrame:
    return get_history_table(conn, "dbo", "CulturesHistory", culture_ids)


def get_champions_history(conn: pyodbc.Connection, culture_ids: List[int]) -> pd.DataFrame:
    candidates = [("dbo", "ChampionsCulturesHistory"), ("dbo", "ChampionsCultureHistory")]
    for schema, table in candidates:
        if table_exists(conn, schema, table):
            return get_history_table(conn, schema, table, culture_ids)
    return pd.DataFrame(columns=["CultureID", "TimeStamp"])


def build_parent_map(edges: pd.DataFrame) -> Dict[int, List[int]]:
    child_to_parents: Dict[int, List[int]] = {}
    if edges.empty:
        return child_to_parents
    for _, r in edges.iterrows():
        p = int(r["ParentCultureID"])
        c = int(r["ChildCultureID"])
        child_to_parents.setdefault(c, []).append(p)
    return child_to_parents


def trace_lineage_single_parent(final_culture_id: int, parent_map: Dict[int, List[int]]) -> List[int]:
    chain = [int(final_culture_id)]
    cur = int(final_culture_id)
    seen = {cur}
    while cur in parent_map and parent_map[cur]:
        parent = int(parent_map[cur][0])
        if parent in seen:
            break
        chain.append(parent)
        seen.add(parent)
        cur = parent
    return chain


def infer_birth_times_from_combined_history(
    culture_hist: pd.DataFrame,
    champ_hist: pd.DataFrame,
) -> Dict[int, pd.Timestamp]:
    frames = []
    if not culture_hist.empty:
        frames.append(culture_hist[["CultureID", "TimeStamp"]])
    if not champ_hist.empty:
        frames.append(champ_hist[["CultureID", "TimeStamp"]])
    if not frames:
        return {}

    comb = pd.concat(frames, ignore_index=True)
    comb = comb.dropna(subset=["CultureID", "TimeStamp"]).copy()
    comb["CultureID"] = comb["CultureID"].astype(int)
    comb["TimeStamp"] = pd.to_datetime(comb["TimeStamp"])

    birth = comb.groupby("CultureID")["TimeStamp"].min().to_dict()
    return {int(k): pd.to_datetime(v) for k, v in birth.items()}


def last_row_at_or_before(df: pd.DataFrame, culture_id: int, t: pd.Timestamp) -> Optional[pd.Series]:
    if df.empty:
        return None
    sub = df[(df["CultureID"] == culture_id) & (df["TimeStamp"] <= t)]
    if sub.empty:
        return None
    return sub.sort_values("TimeStamp").iloc[-1]


def source_row_consistent(
    crow: Optional[pd.Series],
    hrow: Optional[pd.Series],
    t: Optional[pd.Timestamp] = None,
    prefer_culture_history: bool = False,
) -> Optional[pd.Series]:
    # When a row exists exactly at the requested timestamp, use that source first.
    if t is not None:
        t_cmp = pd.to_datetime(t)
        if hrow is not None:
            h_ts = pd.to_datetime(hrow.get("TimeStamp", None), errors="coerce")
            if pd.notna(h_ts) and h_ts == t_cmp:
                return hrow
        if crow is not None:
            c_ts = pd.to_datetime(crow.get("TimeStamp", None), errors="coerce")
            if pd.notna(c_ts) and c_ts == t_cmp:
                return crow

    if prefer_culture_history:
        if hrow is not None:
            return hrow
        if crow is not None:
            return crow
        return None

    if crow is not None:
        return crow
    if hrow is not None:
        return hrow
    return None


def compute_od_from_single_row(row: Optional[pd.Series]) -> Optional[float]:
    if row is None:
        return None

    v = row.get("Converted_OD", None)
    if v is not None and not pd.isna(v):
        return float(v)

    a = row.get("OD_FlEx482Em510", None)
    b = row.get("OD_FlEx587Em611", None)
    if pd.isna(a) and pd.isna(b):
        return None

    a_val = 0.0 if pd.isna(a) else float(a)
    b_val = 0.0 if pd.isna(b) else float(b)
    return a_val + b_val


def get_fluor_from_single_row(row: Optional[pd.Series], col: str) -> Optional[float]:
    if row is None:
        return None
    v = row.get(col, None)
    if v is None or pd.isna(v):
        return None
    return float(v)


def well_sort_key(well: str) -> Tuple[str, int]:
    m = re.match(r"^([A-Za-z]+)(\d+)$", well.strip())
    if not m:
        return (well, 0)
    return (m.group(1).upper(), int(m.group(2)))


def determine_active_plate_id(culture_hist: pd.DataFrame, cultures: pd.DataFrame) -> int:
    if culture_hist.empty:
        raise RuntimeError("CulturesHistory is empty")

    t_max = pd.to_datetime(culture_hist["TimeStamp"]).max()
    at_last = culture_hist[culture_hist["TimeStamp"] == t_max][["CultureID"]].drop_duplicates()

    merged = at_last.merge(cultures[["CultureID", "PlateID"]], on="CultureID", how="left").dropna(subset=["PlateID"])
    if merged.empty:
        raise RuntimeError("Cannot map latest CulturesHistory rows to PlateID")

    plates = sorted(merged["PlateID"].astype(int).unique().tolist())
    active_plate = max(plates)

    log(f"Latest TimeStamp: {t_max}")
    log(f"Candidate plates at latest TimeStamp: {plates}")
    log(f"Active PlateID picked: {active_plate}")

    return active_plate


def determine_active_cultures_by_plate_only(cultures: pd.DataFrame, active_plate: int, n_active: int) -> List[int]:
    plate_cult = cultures[cultures["PlateID"] == active_plate].copy()
    if plate_cult.empty:
        raise RuntimeError(f"No cultures found for active PlateID={active_plate}")

    plate_cult["__well_key__"] = plate_cult["WellID"].map(well_sort_key)
    plate_cult = plate_cult.sort_values("__well_key__")

    ids = plate_cult["CultureID"].astype(int).tolist()

    if n_active <= 0:
        raise RuntimeError("Active culture count from ImportPlatePattern is 0")

    if len(ids) < n_active:
        log(f"WARNING: Plate {active_plate} has only {len(ids)} cultures but n_active={n_active}. Using all.")
        return ids

    return ids[:n_active]


def build_time_index(
    culture_hist: pd.DataFrame,
    champ_hist: pd.DataFrame,
    tracks: List[List[int]],
    birth_times: Dict[int, pd.Timestamp],
) -> List[pd.Timestamp]:
    times: Set[pd.Timestamp] = set()

    if not culture_hist.empty:
        times.update(pd.to_datetime(culture_hist["TimeStamp"].dropna()).tolist())
    if not champ_hist.empty and "TimeStamp" in champ_hist.columns:
        times.update(pd.to_datetime(champ_hist["TimeStamp"].dropna()).tolist())

    for chain in tracks:
        for cid in chain:
            bt = birth_times.get(int(cid))
            if bt is not None:
                times.add(pd.to_datetime(bt))

    return sorted(times, reverse=True)


def choose_current_culture(chain: List[int], birth_times: Dict[int, pd.Timestamp], t: pd.Timestamp) -> Optional[int]:
    current = None
    best_bt = None
    for cid in chain:
        bt = birth_times.get(int(cid))
        if bt is None:
            continue
        if bt <= t and (best_bt is None or bt > best_bt):
            best_bt = bt
            current = int(cid)
    return current


def build_row_for_time_and_ids(
    t: pd.Timestamp,
    current_ids: List[Optional[int]],
    culture_loc: Dict[int, Dict[str, object]],
    culture_hist: pd.DataFrame,
    champ_hist: pd.DataFrame,
    fluor_cols: List[str],
    prefer_culture_history: bool = False,
) -> Dict[str, object]:
    row: Dict[str, object] = {"time": pd.to_datetime(t)}

    for i, current in enumerate(current_ids, start=1):
        row[f"Culture {i} ID"] = current

        loc = {"PlateID": None, "WellID": None}
        if current is not None:
            loc = culture_loc.get(int(current), {"PlateID": None, "WellID": None})

        hrow = last_row_at_or_before(culture_hist, int(current), t) if current is not None else None
        crow = (
            last_row_at_or_before(champ_hist, int(current), t)
            if (current is not None and not champ_hist.empty)
            else None
        )

        src = source_row_consistent(crow, hrow, t=t, prefer_culture_history=prefer_culture_history)

        row[f"Culture {i} OD"] = compute_od_from_single_row(src)
        for fc in fluor_cols:
            row[f"Culture {i} {fc}"] = get_fluor_from_single_row(src, fc)

        row[f"Culture {i} Plate"] = loc.get("PlateID", None)
        row[f"Culture {i} Well"] = loc.get("WellID", None)

    return row


def _row_culture_ids(series: pd.Series, n: int) -> List[Optional[int]]:
    ids: List[Optional[int]] = []
    for i in range(1, n + 1):
        v = series.get(f"Culture {i} ID", None)
        if v is None or pd.isna(v):
            ids.append(None)
        else:
            ids.append(int(v))
    return ids


def expand_propagation_rows_with_parents(
    df_out: pd.DataFrame,
    n: int,
    parent_map: Dict[int, List[int]],
    culture_loc: Dict[int, Dict[str, object]],
    culture_hist: pd.DataFrame,
    champ_hist: pd.DataFrame,
    fluor_cols: List[str],
) -> Tuple[pd.DataFrame, List[Tuple[int, int]]]:
    """
    Duplicates each propagation row by inserting a parent row immediately below it.
    Parent IDs are resolved from Propagation (child -> parent) first, then adjacent-row
    fallback is used if needed. Returns the expanded dataframe and 0-based data-row
    index pairs to merge in the event column (child_row_pos, parent_row_pos).
    """
    if df_out.empty:
        return df_out, []

    expanded_rows: List[Dict[str, object]] = []
    event_merge_pairs: List[Tuple[int, int]] = []

    for r in range(len(df_out)):
        t = pd.to_datetime(df_out.index[r])
        child_series = df_out.iloc[r]

        child_record = {"time": t}
        child_record.update(child_series.to_dict())
        expanded_rows.append(child_record)

        if r >= len(df_out) - 1:
            continue

        if str(child_series.get("event", "")).strip().lower() != "propagation":
            continue

        older_series = df_out.iloc[r + 1]
        child_ids = _row_culture_ids(child_series, n)
        older_ids = _row_culture_ids(older_series, n)

        parent_ids = list(child_ids)
        changed_any = False
        for idx in range(n):
            c_id = child_ids[idx]
            o_id = older_ids[idx]
            if c_id is None:
                continue

            mapped_parent: Optional[int] = None
            parents = parent_map.get(int(c_id), [])
            if parents:
                mapped_parent = int(parents[0])

            if mapped_parent is not None:
                if mapped_parent != c_id:
                    parent_ids[idx] = mapped_parent
                    changed_any = True
                    continue
                # If mapping resolves to self unexpectedly, keep older-row fallback.

            if o_id is None:
                continue
            if c_id != o_id:
                parent_ids[idx] = int(o_id)
                changed_any = True

        if not changed_any:
            continue

        parent_record = build_row_for_time_and_ids(
            t,
            parent_ids,
            culture_loc,
            culture_hist,
            champ_hist,
            fluor_cols,
            prefer_culture_history=True,
        )
        parent_record["event"] = child_series.get("event", "propagation")
        expanded_rows.append(parent_record)

        child_pos = len(expanded_rows) - 2
        parent_pos = len(expanded_rows) - 1
        event_merge_pairs.append((child_pos, parent_pos))

    out = pd.DataFrame(expanded_rows)
    if out.empty:
        return df_out, []

    out["time"] = pd.to_datetime(out["time"])
    out = out.set_index("time")
    out.index.name = df_out.index.name

    for c in df_out.columns:
        if c not in out.columns:
            out[c] = None
    out = out.reindex(columns=df_out.columns)
    return out, event_merge_pairs


def add_cumulative_time_hours_column(df_out: pd.DataFrame, col_name: str = "cumulative Time") -> pd.DataFrame:
    if col_name in df_out.columns:
        df_out = df_out.drop(columns=[col_name])

    if df_out.empty:
        insert_at = 1 if "event" in df_out.columns else 0
        df_out.insert(insert_at, col_name, [])
        return df_out

    times = pd.to_datetime(df_out.index)
    t0 = times.min()
    cumulative_hours = (times - t0).total_seconds() / 3600.0

    insert_at = 1 if "event" in df_out.columns else 0
    df_out.insert(insert_at, col_name, cumulative_hours)
    return df_out


def validate_transitions(
    df_out: pd.DataFrame,
    n: int,
    propagation_edges: Set[Tuple[int, int]],
) -> pd.DataFrame:
    """
    df_out is indexed by time descending.
    For each culture column, check adjacent row transitions:
      if ID changes from old_id at next row (older) to new_id at current row (newer),
      require (old_id, new_id) in Propagation.
    """
    violations: List[Dict[str, object]] = []

    times = df_out.index.tolist()
    if len(times) < 2:
        return pd.DataFrame(columns=["CultureIndex", "NewerTime", "OlderTime", "OlderID", "NewerID", "EdgeExists"])

    for i_col in range(1, n + 1):
        col = f"Culture {i_col} ID"
        if col not in df_out.columns:
            continue

        for r in range(0, len(times) - 1):
            t_new = times[r]
            t_old = times[r + 1]
            new_id = df_out.iloc[r][col]
            old_id = df_out.iloc[r + 1][col]

            if pd.isna(new_id) or pd.isna(old_id):
                continue

            new_id_int = int(new_id)
            old_id_int = int(old_id)

            if new_id_int == old_id_int:
                continue

            edge_ok = (old_id_int, new_id_int) in propagation_edges
            if not edge_ok:
                violations.append(
                    {
                        "CultureIndex": i_col,
                        "NewerTime": t_new,
                        "OlderTime": t_old,
                        "OlderID": old_id_int,
                        "NewerID": new_id_int,
                        "EdgeExists": False,
                    }
                )

    out = pd.DataFrame(violations)
    if not out.empty:
        out["NewerTime"] = pd.to_datetime(out["NewerTime"])
        out["OlderTime"] = pd.to_datetime(out["OlderTime"])
        out = out.sort_values(["CultureIndex", "NewerTime"], ascending=[True, False])
    return out


def add_event_column(df_out: pd.DataFrame, n: int) -> pd.DataFrame:
    """
    Adds a column named 'event' (string) right after the time index when written to Excel.

    event meaning (based on change in any Culture i ID compared to the next older row):
      propagation: at least one Culture i ID changes
      intermediate_fl: no Culture i ID changes
      start: last (oldest) row where no comparison is available
    """
    if df_out.empty:
        df_out.insert(0, "event", [])
        return df_out

    id_cols = [f"Culture {i} ID" for i in range(1, n + 1) if f"Culture {i} ID" in df_out.columns]
    if not id_cols:
        df_out.insert(0, "event", ["start"] * len(df_out))
        return df_out

    events: List[str] = []
    for r in range(len(df_out)):
        if r == len(df_out) - 1:
            events.append("start")
            continue

        cur_row = df_out.iloc[r]
        next_row = df_out.iloc[r + 1]

        changed = False
        for c in id_cols:
            a = cur_row.get(c, None)
            b = next_row.get(c, None)
            if pd.isna(a) or pd.isna(b):
                continue
            if int(a) != int(b):
                changed = True
                break

        events.append("propagation" if changed else "intermediate_fl")

    if "event" in df_out.columns:
        df_out = df_out.drop(columns=["event"])

    df_out.insert(0, "event", events)
    return df_out


def export_xlsx_with_formatting(
    out_xlsx_path: str,
    df_out: pd.DataFrame,
    n: int,
    fluor_cols: List[str],
    event_merge_pairs: List[Tuple[int, int]],
) -> None:
    with pd.ExcelWriter(out_xlsx_path, engine="openpyxl") as writer:
        df_out.to_excel(writer, sheet_name="CultureHistory", index=True)

        ws = writer.sheets["CultureHistory"]

        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        fill_id = PatternFill("solid", fgColor="E2EFDA")
        fill_od = PatternFill("solid", fgColor="FFF2CC")
        fill_flu = PatternFill("solid", fgColor="F8CBAD")
        fill_loc = PatternFill("solid", fgColor="D9E1F2")
        fill_event = PatternFill("solid", fgColor="E7E6E6")
        fill_time = PatternFill("solid", fgColor="FFFFFF")

        header_font = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Freeze: time column (A), event column (B), cumulative Time (C), plus header row
        ws.freeze_panes = "D2"

        max_col = ws.max_column
        max_row = ws.max_row

        for c in range(1, max_col + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font
            cell.alignment = center

        col_types: Dict[int, str] = {}
        for c in range(1, max_col + 1):
            h = ws.cell(row=1, column=c).value
            if c == 1 or str(h).lower() == "time":
                col_types[c] = "time"
                continue

            hs = str(h)
            hs_lower = hs.strip().lower()

            if hs_lower == "event":
                col_types[c] = "event"
            elif hs.endswith(" ID"):
                col_types[c] = "id"
            elif hs.endswith(" OD"):
                col_types[c] = "od"
            elif " FlEx" in hs:
                col_types[c] = "flu"
            elif hs.endswith(" Plate") or hs.endswith(" Well"):
                col_types[c] = "loc"
            else:
                col_types[c] = "time"

        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.alignment = center
                t = col_types.get(c, "time")
                if t == "id":
                    cell.fill = fill_id
                elif t == "od":
                    cell.fill = fill_od
                elif t == "flu":
                    cell.fill = fill_flu
                elif t == "loc":
                    cell.fill = fill_loc
                elif t == "event":
                    cell.fill = fill_event
                else:
                    cell.fill = fill_time

        for r in range(2, max_row + 1):
            ws.cell(row=r, column=1).number_format = "yyyy/mm/dd hh:mm"

        cumulative_col_idx = None
        event_col_idx = None
        for c in range(1, max_col + 1):
            h = ws.cell(row=1, column=c).value
            if h == "cumulative Time":
                cumulative_col_idx = c
            if str(h).strip().lower() == "event":
                event_col_idx = c

        if cumulative_col_idx is not None:
            for r in range(2, max_row + 1):
                ws.cell(row=r, column=cumulative_col_idx).number_format = "0.00"

        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 18
        if max_col >= 3:
            ws.column_dimensions["C"].width = 16
        for c in range(4, max_col + 1):
            ws.column_dimensions[get_column_letter(c)].width = 16

        if event_col_idx is not None:
            for child_pos, parent_pos in event_merge_pairs:
                r1 = child_pos + 2
                r2 = parent_pos + 2
                if r1 < 2 or r2 > max_row or r2 <= r1:
                    continue
                ws.merge_cells(start_row=r1, start_column=event_col_idx, end_row=r2, end_column=event_col_idx)
                ws.cell(row=r1, column=event_col_idx).alignment = center


def main() -> int:
    log(f"START {SCRIPT_NAME}")
    log(f"Args: {sys.argv}")

    if len(sys.argv) < 2:
        log("ERROR: Missing ExperimentID. Usage: Data.exe <ExperimentID>")
        safe_print("Missing ExperimentID")
        return 1

    try:
        experiment_id = int(sys.argv[1])
    except Exception as e:
        log(f"ERROR: ExperimentID not integer: {e}")
        safe_print("ExperimentID must be an integer")
        return 1

    out_name = f"Experiment_{experiment_id}_CultureHistory_{TS}.xlsx"
    out_path = os.path.join(OUT_DIR, out_name)
    log(f"Output target: {out_path}")
    safe_print(f"Output target: {out_path}")

    try:
        with connect() as conn:
            ancestor_plates = get_ancestor_plate_ids(conn, experiment_id)
            if not ancestor_plates:
                log(f"ERROR: No ancestor plates found for ExperimentID={experiment_id}")
                return 1

            plate_ids = get_all_plate_ids(conn, ancestor_plates)
            cultures = get_cultures(conn, plate_ids)
            if cultures.empty:
                log("ERROR: No cultures found on plates in scope")
                return 1

            n_active = get_cells_count_from_import_plate_pattern(conn, plate_ids)
            log(f"Active culture count from ImportPlatePattern Cells: {n_active}")
            if n_active <= 0:
                log("ERROR: Active culture count is 0")
                return 1

            scope_culture_ids = cultures["CultureID"].dropna().astype(int).unique().tolist()

            edges = get_propagation_edges(conn)
            parent_map = build_parent_map(edges)

            culture_hist = get_cultures_history(conn, scope_culture_ids)
            if culture_hist.empty:
                log("ERROR: No CulturesHistory rows found for cultures in scope")
                return 1

            champ_hist = get_champions_history(conn, scope_culture_ids)

            active_plate = determine_active_plate_id(culture_hist, cultures)
            finals = determine_active_cultures_by_plate_only(cultures, active_plate, n_active)

            log(f"Active PlateID: {active_plate}")
            log(f"Final cultures count={len(finals)}")

            tracks = [trace_lineage_single_parent(fc, parent_map) for fc in finals]
            birth_times = infer_birth_times_from_combined_history(culture_hist, champ_hist)
            log(f"Birth times computed: {len(birth_times)}")

            times_sorted = build_time_index(culture_hist, champ_hist, tracks, birth_times)
            log(f"Time index size: {len(times_sorted)}")

            culture_loc = cultures.set_index("CultureID")[["PlateID", "WellID"]].to_dict(orient="index")
            fluor_cols = []
            for c in ["FlEx482Em510", "FlEx587Em611"]:
                if c in culture_hist.columns or c in champ_hist.columns:
                    fluor_cols.append(c)

            rows: List[Dict[str, object]] = []
            for t in times_sorted:
                current_ids = [choose_current_culture(chain, birth_times, t) for chain in tracks]
                row = build_row_for_time_and_ids(t, current_ids, culture_loc, culture_hist, champ_hist, fluor_cols)
                rows.append(row)

            df_out = pd.DataFrame(rows)
            df_out["time"] = pd.to_datetime(df_out["time"])
            df_out = df_out.set_index("time").sort_index(ascending=False)
            df_out.index.name = "time"

            # Add event column (propagation vs intermediate_fl), to be frozen with time
            df_out = add_event_column(df_out, len(finals))

            # Duplicate propagation rows with parent CultureIDs at the same timestamp (child row stays on top)
            df_out, event_merge_pairs = expand_propagation_rows_with_parents(
                df_out, len(finals), parent_map, culture_loc, culture_hist, champ_hist, fluor_cols
            )

            # Add cumulative time in hours since experiment start (oldest timestamp in export)
            df_out = add_cumulative_time_hours_column(df_out, "cumulative Time")

            # Reorder columns so Culture i ID block is at the very end
            ordered_cols: List[str] = []
            ordered_cols.append("event")
            ordered_cols.append("cumulative Time")

            for i in range(1, len(finals) + 1):
                ordered_cols.append(f"Culture {i} OD")

            for fc in fluor_cols:
                for i in range(1, len(finals) + 1):
                    ordered_cols.append(f"Culture {i} {fc}")

            for i in range(1, len(finals) + 1):
                ordered_cols.append(f"Culture {i} Plate")
                ordered_cols.append(f"Culture {i} Well")

            for i in range(1, len(finals) + 1):
                ordered_cols.append(f"Culture {i} ID")

            df_out = df_out.reindex(columns=ordered_cols)

            export_xlsx_with_formatting(out_path, df_out, len(finals), fluor_cols, event_merge_pairs)

        log(f"SAVED: {out_path}")
        safe_print(f"Saved: {out_path}")
        return 0

    except Exception as e:
        log(f"FATAL: {type(e).__name__}: {e}")
        safe_print(f"ERROR: {e}")
        return 1

    finally:
        log(f"END {SCRIPT_NAME}")


if __name__ == "__main__":
    raise SystemExit(main())
